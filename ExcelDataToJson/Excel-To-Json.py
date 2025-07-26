from openpyxl import load_workbook
import json
import argparse

parser = argparse.ArgumentParser(
    prog="ExcelDataToJson",
    description="""Converts excel into json. Must have each sheet start with
    "Accession #	Title	Location	Subject Headings	Description	Date	Created By	Size	Condition	Status	Donated By	Acquisition	Attachment	Notes"
    The name of the sheets is used.""",
)
parser.add_argument("filename", type=str, help="The path of the excel file")
parser.add_argument(
    "-o",
    type=str,
    help="The filepath of the output, must have .json",
    default="out.json",
)
parser.add_argument(
    "--sheet",
    action=argparse.BooleanOptionalAction,
    help="When sheet (default), the artifacts will be divided by their corresponding sheet",
    default=True,
)
args = parser.parse_args()

heading_dict = {
    0: "Accession",
    1: "Title",
    2: "Location",
    3: "SubjectHeadings",
    4: "Description",
    5: "Date",
    6: "CreatedBy",
    7: "Size",
    8: "Condition",
    9: "Status",
    10: "DonatedBy",
    11: "Acquisition",
    12: "Attachment",
    13: "Notes",
}

wb = load_workbook(filename=args.filename)
sheets = wb.sheetnames
db = {sheet: dict() for sheet in sheets} if args.sheet else dict()
for sheet in sheets:
    generator = wb[sheet].values
    for index, row in enumerate([x for x in generator]):
        if row[0] and index > 0:  # only rows with data
            row = row[:14]
            artifact_data = {heading_dict[i]: v for i, v in enumerate(row)}
            if args.sheet:
                db[sheet][artifact_data["Accession"]] = artifact_data
            else:
                db[artifact_data["Accession"]] = artifact_data

with open(args.o, "w") as f:
    f.write(json.dumps(db, indent=2))

print("Saved to", args.o)
